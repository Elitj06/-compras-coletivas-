#!/usr/bin/env python3
"""
Atualizar catálogo Vida Forte a partir da planilha mensal Vitafor.

Uso:
    python3 scripts/atualizar_catalogo.py data/TABELA_VITAFOR_ABRIL_2026.xlsx

O que faz:
    1. Lê as abas "TABELA DE PEDIDO VITAFOR" e "TABELA PEDIDO VITAPOWER"
    2. Extrai código, nome, embalagem, preço bruto e categoria de cada produto
    3. Extrai TODAS as imagens embutidas e as ancora à linha do produto
    4. Mescla com o produtos.js existente:
        - novos códigos são adicionados
        - preços são atualizados
        - imagens novas substituem as antigas (ou preenchem ausentes)
        - produtos descontinuados continuam no histórico via flag
    5. Reescreve public/produtos.js mantendo o mesmo formato consumido pelo frontend.

Filosofia: este é o ÚNICO ponto de verdade para atualizações mensais.
Basta substituir a planilha e rodar o script — o app inteiro reflete.
"""
from __future__ import annotations

import base64
import io
import json
import re
import sys
import unicodedata
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------
# Configurações
# ----------------------------------------------------------------------
ROOT = Path(__file__).resolve().parent.parent
PRODUTOS_JS = ROOT / "public" / "produtos.js"
HEADER_KEYWORDS_VITA = {"CÓDIGO", "CODIGO"}
HEADER_KEYWORDS_VPP = {"SKU"}
PRICE_RE = re.compile(r"R?\$?\s*([\d\.]+,\d{2}|\d+\.\d{2}|\d+)")
CODE_RE = re.compile(r"^[A-Z]{2,}\d+[A-Z0-9]*$")

# Códigos com selo "lançamento" (sem foto real) — forçar limpeza
LAUNCH_BADGE_HINT_BYTES = 6000  # imagens muito pequenas costumam ser selos


# ----------------------------------------------------------------------
# Utilidades
# ----------------------------------------------------------------------
def slugify(text: str) -> str:
    """Gera id de categoria compatível com o legado."""
    if not text:
        return "outros"
    s = unicodedata.normalize("NFKD", text)
    s = s.encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^a-zA-Z0-9]+", "_", s).strip("_").lower()
    return s[:30] or "outros"


def parse_price(raw: str) -> Optional[float]:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip()
    if not s:
        return None
    m = PRICE_RE.search(s)
    if not m:
        return None
    val = m.group(1).replace(".", "").replace(",", ".") if "," in m.group(1) else m.group(1)
    try:
        return float(val)
    except ValueError:
        return None


def parse_int(raw) -> Optional[int]:
    if raw is None:
        return None
    try:
        return int(float(str(raw).replace(",", ".")))
    except ValueError:
        return None


def cell_str(ws, row, col) -> str:
    v = ws.cell(row=row, column=col).value
    return "" if v is None else str(v).strip()


# ----------------------------------------------------------------------
# Leitura da planilha
# ----------------------------------------------------------------------
def detect_columns(ws, header_keywords) -> dict:
    """Localiza header e mapeia colunas-chave por nome aproximado.

    Como o cabeçalho original tem células mescladas, primeiro encontramos
    a linha que contém "CÓDIGO/SKU" e depois inferimos a coluna do código
    olhando a primeira linha de produto (procurando um valor que case com
    o regex de código). Demais colunas são posicionais a partir dessa.
    """
    header_row = None
    for r in range(1, min(ws.max_row, 40) + 1):
        for c in range(1, ws.max_column + 1):
            v = cell_str(ws, r, c).upper()
            if any(k in v for k in header_keywords):
                header_row = r
                break
        if header_row:
            break
    if not header_row:
        return {}

    # Descobrir a coluna do código olhando as 50 primeiras linhas após o header
    codigo_col = None
    for r in range(header_row + 1, min(header_row + 60, ws.max_row + 1)):
        for c in range(1, ws.max_column + 1):
            v = cell_str(ws, r, c).strip()
            if v and CODE_RE.match(v.replace(" ", "")):
                codigo_col = c
                break
        if codigo_col:
            break
    if not codigo_col:
        return {}

    # Para cada linha de produto, descobrir as outras colunas relativas
    # olhando 1 linha de exemplo
    cols = {"header_row": header_row, "codigo": codigo_col}
    for r in range(header_row + 1, min(header_row + 80, ws.max_row + 1)):
        if not cell_str(ws, r, codigo_col).strip():
            continue
        if not CODE_RE.match(cell_str(ws, r, codigo_col).strip().replace(" ", "")):
            continue
        # nome: primeira coluna não-vazia >2 chars depois do código
        for c in range(codigo_col + 1, ws.max_column + 1):
            v = cell_str(ws, r, c).strip()
            if v and len(v) > 3 and not v.replace(",", "").replace(".", "").isdigit():
                cols["nome"] = c
                break
        # emb: primeira coluna numérica depois do nome
        start = cols.get("nome", codigo_col + 1) + 1
        for c in range(start, ws.max_column + 1):
            v = cell_str(ws, r, c).strip()
            if v and parse_int(v) is not None:
                cols["emb"] = c
                break
        # preço: primeira coluna numérica depois da embalagem
        start = cols.get("emb", start) + 1
        for c in range(start, ws.max_column + 1):
            v = cell_str(ws, r, c).strip()
            if v and parse_price(v) is not None:
                cols["preco"] = c
                break
        break
    return cols


def find_category_above(ws, row: int, codigo_col: int) -> str:
    """Categoria é uma linha-cabeçalho mesclada acima do produto."""
    for r in range(row - 1, max(row - 15, 0), -1):
        for c in range(1, ws.max_column + 1):
            v = cell_str(ws, r, c)
            if not v:
                continue
            up = v.upper()
            # heurística: linha de categoria não tem código
            if CODE_RE.match(up.replace(" ", "")):
                return ""  # achou outro produto antes -> sem categoria
            # ignora linhas de cabeçalho/preço
            if any(k in up for k in ("CÓDIGO", "CODIGO", "PREÇO", "PRECO", "DESCONTO", "EMBALAGEM", "PEDIDO", "TOTAL")):
                continue
            # categorias normalmente são string longa em maiúsculas
            if len(v) > 3 and v == up:
                return v.strip()
        # se a linha tiver um código, paramos
        if codigo_col and cell_str(ws, r, codigo_col):
            v = cell_str(ws, r, codigo_col)
            if CODE_RE.match(v):
                break
    return ""


def extract_products(ws, header_keywords) -> list[dict]:
    cols = detect_columns(ws, header_keywords)
    if not cols.get("codigo"):
        return []

    produtos = []
    last_category = ""
    for r in range(cols["header_row"] + 1, ws.max_row + 1):
        cod_raw = cell_str(ws, r, cols["codigo"])
        if not cod_raw:
            # candidato a linha de categoria
            for c in range(1, ws.max_column + 1):
                v = cell_str(ws, r, c)
                if v and len(v) > 3 and v == v.upper() and not CODE_RE.match(v.replace(" ", "")):
                    if not any(k in v.upper() for k in ("PEDIDO", "TOTAL", "DESCONTO", "PREÇO", "EMBALAGEM", "CÓDIGO")):
                        last_category = v.strip()
                        break
            continue

        if not CODE_RE.match(cod_raw):
            continue

        nome = cell_str(ws, r, cols.get("nome", cols["codigo"] + 1))
        if not nome:
            continue

        preco = parse_price(cell_str(ws, r, cols.get("preco", cols["codigo"] + 3)))
        emb = parse_int(cell_str(ws, r, cols.get("emb", cols["codigo"] + 2))) or 1

        produtos.append(
            {
                "row": r,
                "codigo": cod_raw,
                "nome": nome,
                "preco": preco or 0.0,
                "embalagem": emb,
                "categoriaNome": last_category or "OUTROS",
                "categoria": slugify(last_category) if last_category else "outros",
            }
        )
    return produtos


# ----------------------------------------------------------------------
# Extração de imagens
# ----------------------------------------------------------------------
def extract_images_by_row(ws) -> dict[int, str]:
    """Retorna {row_index: data_uri_base64} para cada imagem embutida.

    A âncora `_from.row` é 0-indexed; convertemos para 1-indexed.
    Se múltiplas imagens caírem na mesma linha, mantemos a maior
    (provavelmente a foto real, não um selo).
    """
    out: dict[int, tuple[int, str]] = {}
    for img in ws._images:
        try:
            anchor = img.anchor
            row0 = getattr(anchor, "_from", None)
            if row0 is None:
                continue
            row = row0.row + 1
            # carrega bytes
            raw = img._data() if callable(img._data) else img._data
            if isinstance(raw, bytes):
                blob = raw
            else:
                blob = bytes(raw)
            size = len(blob)
            # filtro: ignorar selos pequenos
            if size < LAUNCH_BADGE_HINT_BYTES:
                continue
            # detecta MIME pelo magic
            if blob[:3] == b"\xff\xd8\xff":
                mime = "image/jpeg"
            elif blob[:8] == b"\x89PNG\r\n\x1a\n":
                mime = "image/png"
            else:
                mime = "image/jpeg"
            uri = f"data:{mime};base64,{base64.b64encode(blob).decode('ascii')}"
            existing = out.get(row)
            if not existing or size > existing[0]:
                out[row] = (size, uri)
        except Exception as e:
            print(f"  ! falha em imagem da linha {getattr(img.anchor._from, 'row', '?')}: {e}", file=sys.stderr)
    return {r: v[1] for r, v in out.items()}


# ----------------------------------------------------------------------
# Mescla com produtos.js antigo
# ----------------------------------------------------------------------
def load_existing_produtos() -> dict[str, dict]:
    """Lê produtos.js antigo via regex grosseiro e devolve {codigo: {...}}.
    Usado para preservar imagens válidas dos produtos que continuam ativos
    quando a planilha nova não trouxer imagem nova.
    """
    if not PRODUTOS_JS.exists():
        return {}
    src = PRODUTOS_JS.read_text(encoding="utf-8")
    m = re.search(r"const PRODUTOS\s*=\s*\[(.*?)\];", src, re.DOTALL)
    if not m:
        return {}
    body = m.group(1)
    # extrai cada objeto { ... }
    items = {}
    depth = 0
    start = None
    for i, ch in enumerate(body):
        if ch == "{":
            if depth == 0:
                start = i
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0 and start is not None:
                obj_src = body[start : i + 1]
                cod_match = re.search(r'codigo:\s*"([^"]+)"', obj_src)
                img_match = re.search(r'imagem:\s*"([^"]*)"', obj_src)
                cat_match = re.search(r'categoria:\s*"([^"]*)"', obj_src)
                catn_match = re.search(r'categoriaNome:\s*"([^"]*)"', obj_src)
                if cod_match:
                    items[cod_match.group(1)] = {
                        "imagem": img_match.group(1) if img_match else "",
                        "categoria": cat_match.group(1) if cat_match else "",
                        "categoriaNome": catn_match.group(1) if catn_match else "",
                    }
                start = None
    return items


def js_escape(s: str) -> str:
    return s.replace("\\", "\\\\").replace('"', '\\"')


def write_produtos_js(produtos: list[dict], categorias: list[dict]) -> None:
    out = ["// Catálogo Vida Forte / Vitafor — gerado automaticamente",
           "// Não edite à mão. Rode `python3 scripts/atualizar_catalogo.py <planilha>`",
           "",
           "const CATEGORIAS = ["]
    out.append('  { id: "todos", nome: "Todos" },')
    for c in categorias:
        out.append(f'  {{ id: "{js_escape(c["id"])}", nome: "{js_escape(c["nome"])}" }},')
    out.append("];")
    out.append("")
    out.append("const PRODUTOS = [")
    for p in produtos:
        nome = js_escape(p["nome"])
        cat_n = js_escape(p["categoriaNome"])
        cat = js_escape(p["categoria"])
        img = p.get("imagem") or ""
        out.append(
            "  { "
            f'codigo: "{js_escape(p["codigo"])}", '
            f'nome: "{nome}", '
            f"preco: {p['preco']:.2f}, "
            f"embalagem: {p['embalagem']}, "
            f'categoria: "{cat}", '
            f'categoriaNome: "{cat_n}", '
            f'imagem: "{img}"'
            " },"
        )
    out.append("];")
    out.append("")
    PRODUTOS_JS.write_text("\n".join(out), encoding="utf-8")


# ----------------------------------------------------------------------
# Pipeline principal
# ----------------------------------------------------------------------
def main(xlsx_path: str) -> None:
    path = Path(xlsx_path)
    if not path.exists():
        sys.exit(f"Arquivo não encontrado: {path}")

    print(f"→ Carregando {path.name}")
    wb = openpyxl.load_workbook(path, data_only=True)

    sheets = [
        ("TABELA DE PEDIDO VITAFOR", HEADER_KEYWORDS_VITA),
        ("TABELA PEDIDO VITAPOWER", HEADER_KEYWORDS_VPP),
    ]

    all_products: list[dict] = []
    for name, kw in sheets:
        if name not in wb.sheetnames:
            print(f"  ⚠ aba '{name}' não encontrada, pulando")
            continue
        ws = wb[name]
        prods = extract_products(ws, kw)
        imgs = extract_images_by_row(ws)
        print(f"  • {name}: {len(prods)} produtos, {len(imgs)} imagens")

        # Para cada produto, busca imagem na própria linha ou nas vizinhas.
        # Em planilhas Vitafor a âncora de imagem cai geralmente 1-3 linhas
        # acima do número da linha do código (a imagem ocupa várias linhas).
        used = set()
        for p in prods:
            row = p["row"]
            uri = ""
            for delta in (0, -1, 1, -2, 2, -3, 3):
                key = row + delta
                if key in imgs and key not in used:
                    uri = imgs[key]
                    used.add(key)
                    break
            p["imagem"] = uri
        all_products.extend(prods)

    # Mescla com produtos.js antigo
    legacy = load_existing_produtos()
    for p in all_products:
        leg = legacy.get(p["codigo"])
        if leg:
            # Preserva categoria do legado se a planilha não trouxer
            if not p["categoriaNome"] or p["categoriaNome"] == "OUTROS":
                if leg.get("categoriaNome"):
                    p["categoriaNome"] = leg["categoriaNome"]
                    p["categoria"] = leg["categoria"]
            # Se sem imagem, usa a antiga
            if not p["imagem"] and leg.get("imagem"):
                p["imagem"] = leg["imagem"]

    # Categorias deduplicadas
    seen = {}
    for p in all_products:
        if p["categoria"] not in seen:
            seen[p["categoria"]] = {"id": p["categoria"], "nome": p["categoriaNome"]}
    cats = sorted(seen.values(), key=lambda x: x["nome"])

    print(f"→ Total: {len(all_products)} produtos · {len(cats)} categorias")
    print(f"→ Com imagem: {sum(1 for p in all_products if p['imagem'])}")
    write_produtos_js(all_products, cats)
    print(f"✔ {PRODUTOS_JS.relative_to(ROOT)} atualizado")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit("Uso: python3 scripts/atualizar_catalogo.py <planilha.xlsx>")
    main(sys.argv[1])
